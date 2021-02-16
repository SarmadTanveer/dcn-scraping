from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import json
def createSheet(fileName, sheetName, headings):
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    dest_filename = os.path.join(THIS_FOLDER, fileName)

    wb = load_workbook(dest_filename)
    ws = wb.create_sheet(sheetName)
    ws.append(headings)

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CCFFFF', fill_type='solid')

    wb.save(filename=dest_filename)
    wb.close()

def saveData(fileName, sheetName, data):
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    dest_filename = os.path.join(THIS_FOLDER, fileName)

    wb = load_workbook(dest_filename)
    ws = wb[sheetName]
    ws.append(data)
    wb.save(filename=dest_filename)

